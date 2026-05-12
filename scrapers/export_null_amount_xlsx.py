"""
Generate an Excel workbook listing every deal in Supabase that still has
amount_eur IS NULL. Useful for manually filling in amounts that the
scraper / re-extraction couldn't recover (e.g. via a public PitchBook
record, the company's own press release, follow-up reporting, etc).

Writes to data/null_amount_deals.xlsx. The frontend doesn't load this
file — it's just an output artifact published as a static asset so it's
downloadable from the live site at /data/null_amount_deals.xlsx.
"""

import os
import sys
from urllib.parse import urlparse

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_KEY"]

OUT_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "null_amount_deals.xlsx")
OUT_PATH = os.path.abspath(OUT_PATH)

COLUMNS = [
    ("Company",        "company",         32),
    ("Country",        "country",         20),
    ("Sector",         "sector",          14),
    ("Stage",          "stage",           12),
    ("Year",           "year",             7),
    ("Quarter",        "quarter",          9),
    ("Lead investor",  "lead_investor",   30),
    ("Description",    "description",     60),
    ("Source URL",     "_url",            70),
    ("Deal ID",        "id",              36),
]


def fetch_all():
    rows = []
    offset = 0
    page = 1000
    while True:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/deals"
            f"?amount_eur=is.null"
            f"&select=id,company,country,sector,stage,year,quarter,lead_investor,description,source_urls"
            f"&order=year.desc,quarter.desc,company.asc"
            f"&limit={page}&offset={offset}",
            headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"},
            timeout=30,
        )
        r.raise_for_status()
        batch = r.json()
        rows.extend(batch)
        if len(batch) < page:
            break
        offset += page
    return rows


def main():
    rows = fetch_all()
    print(f"Found {len(rows)} NULL-amount deals.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Null amount deals"

    # Header
    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_idx, (heading, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=heading)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    # Body
    for r_idx, deal in enumerate(rows, start=2):
        # First URL only — most rows have just one anyway.
        url = (deal.get("source_urls") or [None])[0]
        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            if key == "_url":
                if url:
                    cell = ws.cell(row=r_idx, column=col_idx, value=url)
                    cell.hyperlink = url
                    cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                else:
                    ws.cell(row=r_idx, column=col_idx, value="")
            else:
                v = deal.get(key)
                ws.cell(row=r_idx, column=col_idx, value=v if v is not None else "")
        ws.row_dimensions[r_idx].height = 18

    # Freeze the header.
    ws.freeze_panes = "A2"
    # Autofilter across the full table.
    ws.auto_filter.ref = ws.dimensions

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    size_kb = os.path.getsize(OUT_PATH) / 1024
    print(f"Wrote {OUT_PATH} ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
